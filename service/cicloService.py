from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import date, datetime
from models.recetaxciclo import RecetaXCiclo
from models.ciclo import Ciclo
from models.recetaxciclo import Recetario
from models.torre import Torre

from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from io import BytesIO

import logging

logger = logging.getLogger("uvicorn")
### -------------------------DATOS CONSERVAR ------------------------
def buscarCiclos1(idReceta, listaRecetaXCiclo, listaReceta_dic, listaCiclos_dic):
    return [
        {
            "id_ciclo": recetaXCiclo.id_ciclo,  # Acceder a recetaXCiclo.id_ciclo
            "pesoTotal": listaCiclos_dic[recetaXCiclo.id_ciclo].pesoDesmoldado,
            "tiempoTotal": listaCiclos_dic[recetaXCiclo.id_ciclo].tiempoDesmolde
        }
        for _, recetaXCiclo, _ in listaRecetaXCiclo  # Desempaquetar la tupla correctamente
        if recetaXCiclo.id_recetario == idReceta
    ]

def resumenDeProductividad(db, fecha_inicio:date, fecha_fin:date):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

    respuestaProductividad = {}
    productosRealizados = {}
    totalPeso = 0
    cantidadCiclosTotal = 0

    tablaCiclos = (
        db.query(Ciclo, RecetaXCiclo, Recetario)
        .join(RecetaXCiclo, Ciclo.id == RecetaXCiclo.id_ciclo)
        .join(Recetario, RecetaXCiclo.id_recetario == Recetario.id)
        .filter(Ciclo.fecha_fin.between(fecha_inicio,fecha_fin))
        .all()
    )

    for ciclo, recetaXCiclo, receta in tablaCiclos:
        totalPeso += ciclo.pesoDesmoldado
        cantidadCiclosTotal += 1
        if receta.id not in productosRealizados:
            listaBuscarCiclo = buscarCiclos1(receta.id, tablaCiclos, {r.id: r for _,_, r in tablaCiclos}, {c.id: c for c, _, _ in tablaCiclos})

            pesoFinal = sum(cicloData["pesoTotal"] for cicloData in listaBuscarCiclo)
            tiempoTotalCiclo = sum(cicloData["tiempoTotal"] for cicloData in listaBuscarCiclo)

            productosRealizados[receta.id] = {
                "id_recetario": receta.id,
                "NombreProducto": receta.codigoProducto,
                "pesoTotal": pesoFinal,
                "cantidadCiclos": len(listaBuscarCiclo),
                "tiempoTotal": tiempoTotalCiclo,
            }
    respuestaProductividad["CantidadCiclosCorrectos"] = cantidadCiclosTotal
    respuestaProductividad["PesoTotalCiclos"] = totalPeso / 1000 # Total en Toneladas
    respuestaProductividad["ProductosRealizados"] = list(productosRealizados.values())

    return respuestaProductividad

def generarDocumentoXLMSProductividad(db, fecha_inicio:date, fecha_fin:date):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())
    
    tablaBaseDatos = (
        db.query(Ciclo, RecetaXCiclo, Recetario, Torre)
        .join(RecetaXCiclo, Ciclo.id == RecetaXCiclo.id_ciclo)
        .join(Recetario, RecetaXCiclo.id_recetario == Recetario.id)
        .join(Torre, Ciclo.id_torre == Torre.id)
        .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
        .all()
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Productividad"
    logoPath = "cremona.png"
    img = Image(logoPath)
    img.width = 280
    img.height = 70
    sheet.add_image(img, 'D1')

    sheet.append(["LISTA PRODUCTOS"])
    producto_cell = sheet.cell(row=sheet.max_row, column=1)
    producto_cell.font = Font(bold= True, size=20)

    sheet.append(["Fecha Inicio:", fecha_inicio.strftime("%Y-%m-%d")])
    fechaInicio_cell = sheet.cell(row=sheet.max_row, column=1)
    fechaInicio_cell.font = Font(bold=True, size=12)
    sheet.append(["Fecha Fin:", fecha_fin.strftime("%Y-%m-%d")])
    fechaFin_cell = sheet.cell(row=sheet.max_row, column=1)
    fechaFin_cell.font = Font(bold=True, size=12)

    headers = ["id_recetario", "CodigoProducto","id_etapa","Cantidad Ciclos","PesoTotalDesmoldado","TiempoTotalDesmoldado","NivelesTorre", "NivelesDesmoldadoCorrectamente"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="145f82", end_color="145f82", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    start_row = sheet.max_row + 1
    resultado = []
    productosRealizados = {}

    
    def sumarDatosCiclos(id_recetario, recetaXCiclo, torre_dic, ciclo_dic):
        return [
            [
                ciclo_dic[recetaCiclo.id_ciclo].pesoDesmoldado, 
                ciclo_dic[recetaCiclo.id_ciclo].tiempoDesmolde,
                torre_dic[ciclo_dic[recetaCiclo.id_ciclo].id_torre].cantidadNiveles,
                recetaCiclo.cantidadNivelesFinalizado,
            ] 
            for _, recetaCiclo,_,_ in recetaXCiclo
            if recetaCiclo.id_recetario == id_recetario
        ]

    for ciclo, recetaXCiclo, receta, torre in tablaBaseDatos:
        if receta.id not in productosRealizados:
            productosRealizados[receta.id] = receta.id
            
            resultadoCiclo = sumarDatosCiclos(
                receta.id,
                tablaBaseDatos,
                {r.id: r for _, _, _, r in tablaBaseDatos},
                {c.id: c for c, _, _, _ in tablaBaseDatos}
            )
            print(f"Resultado: Print {resultadoCiclo}")

            # Inicializa una fila con los datos básicos
            fila = [
                receta.id,
                receta.codigoProducto,
                ciclo.id_etapa,
                len(resultadoCiclo)  # Cantidad de ciclos
            ]
            
            # Calcula los totales de cada columna adicional
            if resultadoCiclo:
                sumarCiclos = [0] * len(resultadoCiclo[0])
                for vector in resultadoCiclo:
                    #sumarCiclos = [x + y for x, y in zip(sumarCiclos, vector)]
                    sumarCiclos = [x + y for x, y in zip(sumarCiclos, (v if v is not None else 0 for v in vector))]
                    
                fila.extend(sumarCiclos)
            else:
                fila.extend([0, 0, 0, 0])  # Rellena con ceros si no hay datos
            
            # Agrega la fila completa a `resultado`
            resultado.append(fila)
    for receta in resultado:
        if isinstance(receta, (list, tuple)):
            sheet.append(receta)
        else:
            print(f"Fila inválida: {receta}")


    
    end_row = sheet.max_row
    start_col = 1
    end_col = len(headers)
    table_range = f"{sheet.cell(row=start_row -1, column=start_col).coordinate}:{sheet.cell(row=end_row, column=end_col).coordinate}"
    table_nombre = "ReporteProductividad"
    tabla = Table(displayName=table_nombre, ref=table_range)
    style = TableStyleInfo(showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla.tableStyleInfo = style

    # Agregar la tabla a la hoja
    sheet.add_table(tabla)
    sheet.append([])

    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    workbook.close() 
    excel_stream.seek(0)  
    return excel_stream

def graficosHistoricos(db, fecha_inicio:date, fecha_fin:date):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

    resultado = {}
    listaPeso = []
    listaCiclos = []
    listaProductos = {}

    tablaBaseDatos = (
        db.query(Ciclo, RecetaXCiclo, Recetario)
        .join(RecetaXCiclo, Ciclo.id == RecetaXCiclo.id_ciclo)
        .join(Recetario, RecetaXCiclo.id_recetario == Recetario.id)
        .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
        .all()
    )
    
    def buscarCiclos(idReceta, tablaDatos):
        return [{
            "idCiclo": ciclo.id,
            "pesoDesmontado": receta.pesoPorNivel * recetaXCiclo.cantidadNivelesFinalizado,
            "fecha_fin" : ciclo.fecha_fin.timestamp(),            
        } for ciclo, recetaXCiclo, receta in tablaDatos if idReceta == recetaXCiclo.id_recetario ]

    for ciclo, recetaXCiclo, receta in tablaBaseDatos:
        listaPeso.append({
            "fecha_fin": ciclo.fecha_fin.strftime("%Y-%m-%d"),
            "PesoDiarioProducto": receta.pesoPorNivel * recetaXCiclo.cantidadNivelesFinalizado
        })
        if recetaXCiclo.id_recetario not in listaProductos:
            listaProductos[recetaXCiclo.id_recetario] = {
                "NombreProducto": receta.codigoProducto,
                "ListaDeCiclos": buscarCiclos(recetaXCiclo.id_recetario, tablaBaseDatos)
            }

    # AGRUPAR POR DIA
    def agruparPorDia(datos):
        grouped_by_day = defaultdict(list)
        
        for item in datos:
            fechaFin = item.fecha_fin
            dia = fechaFin.strftime("%Y-%m-%d")
            grouped_by_day[dia].append(item)
        return dict(grouped_by_day)

    listaXDia = agruparPorDia({c for c,_,_ in tablaBaseDatos})
    for clave, valor in listaXDia.items():
        elemento = {}
        elemento["fecha_fin"]= clave
        elemento["CantidadCiclos"]= len(valor)
        listaCiclos.append(elemento)
    
    listaCiclos = sorted(listaCiclos, key=lambda x: datetime.strptime(x['fecha_fin'], "%Y-%m-%d"))
    
    resultado["ListaCiclosDiario"] = listaCiclos
    resultado["pesoTotalDiario "] = listaPeso
    resultado["listaProductos"] = list(listaProductos.values())

    return resultado

def generarDocumentoXLMSGraficos(db, fecha_inicio:date, fecha_fin:date):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

    tabalaDatos = (
        db.query(Ciclo, RecetaXCiclo, Recetario)
        .join(Ciclo, RecetaXCiclo.id_ciclo == Ciclo.id)
        .join(Recetario, RecetaXCiclo.id_recetario == Recetario.id)
        .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
        .all()
    )
    resultado = []

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Graficos listaProductos"

    logoPath = "cremona.png"
    img = Image(logoPath)
    img.width = 280
    img.height = 70
    sheet.add_image(img, 'D1')

    sheet.append(["REPORTE: LISTA PRODUCTOS"])
    producto_cell = sheet.cell(row=sheet.max_row, column=1)
    producto_cell.font = Font(bold= True, size=20)

    sheet.append(["Fecha Inicio:", fecha_inicio.strftime("%Y-%m-%d")])
    fechaInicio_cell = sheet.cell(row=sheet.max_row, column=1)
    fechaInicio_cell.font = Font(bold=True, size=12)
    sheet.append(["Fecha Fin:", fecha_fin.strftime("%Y-%m-%d")])
    fechaFin_cell = sheet.cell(row=sheet.max_row, column=1)
    fechaFin_cell.font = Font(bold=True, size=12)

    headers = ["IdRecetario", "Codigo Producto", "Estado Maquina","IdCiclo","BandaDesmolde","NumeroGripper","Lote", "PesoDesmoldado (KG)","Tiempo total desmolde (minutos)","FechaInicio", "FechaFin"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="145f82", end_color="145f82", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    start_row = sheet.max_row + 1

    for ciclo, recetaXCiclo, receta in tabalaDatos:
        resultado.append([
            recetaXCiclo.id_recetario,
            receta.codigoProducto, 
            ciclo.estadoMaquina,
            ciclo.id, 
            ciclo.bandaDesmolde, 
            receta.nroGripper,
            ciclo.lote,
            ciclo.pesoDesmoldado,
            ciclo.tiempoDesmolde,
            ciclo.fecha_inicio,
            ciclo.fecha_fin, 
        ])

    for receta in resultado:
        sheet.append(receta)

    end_row = sheet.max_row
    start_col = 1
    end_col = len(headers)
    table_range = f"{sheet.cell(row=start_row -1, column=start_col).coordinate}:{sheet.cell(row=end_row, column=end_col).coordinate}"
    table_nombre = "GraficoPoductividad"
    tabla = Table(displayName=table_nombre, ref=table_range)
    style = TableStyleInfo(showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla.tableStyleInfo = style

    # Agregar la tabla a la hoja
    sheet.add_table(tabla)
    sheet.append([])

    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    workbook.close() 
    excel_stream.seek(0)  
    return excel_stream

def get_lista_total_ciclos_productos(db, fecha_inicio:date, fecha_fin:date):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

    print(f"Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}")

    listaResultado = {}

    tablaBDD = (
        db.query(Ciclo, RecetaXCiclo, Recetario)
        .join(RecetaXCiclo, Ciclo.id == RecetaXCiclo.id_ciclo)
        .join(Recetario, RecetaXCiclo.id_recetario == Recetario.id)
        .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
        .all()
    )
    registro = defaultdict(lambda: {"fecha_fin":"","PesoDiarioProducto":0})
    listaPeso = []

    for ciclo, recetaXCiclo, receta in tablaBDD:
        fecha = ciclo.fecha_fin.strftime("%Y-%m-%d")
        peso = ciclo.pesoDesmoldado

        # Si la receta ya está agregada, sumamos el peso
        registro = next((item for item in listaPeso if item["fecha_fin"] == fecha), None)
        if registro:
            registro["PesoDiarioProducto"] += peso
        else:
            listaPeso.append({"fecha_fin": fecha, "PesoDiarioProducto": peso})
    listaResultado["pesoProducto"] = listaPeso

    grouped_by_day = defaultdict(int)

    for ciclo, recetaXCiclo, receta in tablaBDD:
        dia = ciclo.fecha_fin.strftime("%Y-%m-%d")
        grouped_by_day[dia] += 1

    listaResultado["ciclos"] = [
        {"fecha_fin": fecha, "CiclosCompletados": count} 
        for fecha, count in grouped_by_day.items()
    ]

    return listaResultado

def buscarCiclos(id_receta, tabla_datos):
    return [{
        "idCiclo": ciclo.id,
        "pesoDesmontado": ciclo.pesoDesmoldado,
        "fecha_fin": ciclo.fecha_fin.timestamp(),
    } for ciclo, recetaXCiclo, receta in tabla_datos if id_receta == recetaXCiclo.id_recetario]
def get_lista_productos(db, fecha_inicio: date, fecha_fin: date):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

    listaProductos = {}
    tablaBDD = (
        db.query(Ciclo, RecetaXCiclo, Recetario)
        .join(RecetaXCiclo, Ciclo.id == RecetaXCiclo.id_ciclo)
        .join(Recetario, RecetaXCiclo.id_recetario == Recetario.id)
        .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
        .all()
    )
    for ciclo, recetaXCiclo, receta in tablaBDD:
        if recetaXCiclo.id_recetario not in listaProductos:
            listaProductos[recetaXCiclo.id_recetario] = {
                "id_recetario": receta.id,
                "NombreProducto": receta.codigoProducto,
                "ListaDeCiclos": buscarCiclos(recetaXCiclo.id_recetario, tablaBDD)
            }
    return list(listaProductos.values())